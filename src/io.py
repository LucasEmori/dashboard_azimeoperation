"""Carregamento de planilhas Excel com normalizacao de colunas."""
from __future__ import annotations

import glob
import logging
import os
import re
import unicodedata
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import load_workbook

from . import config

log = logging.getLogger("dashboard.io")

EXCEL_EPOCH = datetime(1899, 12, 30)


# ---------------------------------------------------------------------------
# Utilitarios
# ---------------------------------------------------------------------------
def asciize(s) -> str:
    """Converte string Unicode para ASCII removendo acentos."""
    if not isinstance(s, str):
        s = str(s)
    norm = unicodedata.normalize("NFKD", s)
    return norm.encode("ascii", "ignore").decode("ascii")


def col_map(df: pd.DataFrame) -> dict[str, str]:
    """Mapeia nome ASCII -> nome original das colunas do DataFrame."""
    return {asciize(c): c for c in df.columns}


def norm_nf(val) -> int | None:
    """Extrai apenas digitos de um numero de NF (remove letras, espacos, zeros a esquerda)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    digits = re.sub(r"\D", "", str(val))
    return int(digits) if digits else None


def find_file(company: str, prefix: str) -> str:
    """Encontra o arquivo .xlsx na pasta da empresa com o prefixo dado."""
    pasta = config.COMPANY_DIRS[company]
    matches = sorted(glob.glob(os.path.join(str(pasta), f"{prefix}*.xlsx")))
    return matches[0] if matches else None


def _use_bq() -> bool:
    return config.DATA_SOURCE == "bq"


# ---------------------------------------------------------------------------
# Planilha 1 - Produtos Lancados
# ---------------------------------------------------------------------------
def load_p1(company: str) -> pd.DataFrame:
    if _use_bq():
        from . import bq
        return bq.load_p1_bq(company)
    path = find_file(company, config.BASE_DIR.name and "1")
    if path is None:
        log.warning("P1 nao encontrada para %s", company)
        return pd.DataFrame()
    log.info("Lendo P1 (%s): %s", company, os.path.basename(path))
    df = pd.read_excel(path, sheet_name="Plan1", engine="openpyxl")
    cm = col_map(df)
    for key in ("Data", "Data Lancamento", "Data Virada"):
        orig = cm.get(key)
        if orig:
            df[orig] = pd.to_datetime(df[orig], errors="coerce")
    return df


# ---------------------------------------------------------------------------
# Planilha 3 - Notas de Entrada
# ---------------------------------------------------------------------------
def load_p3(company: str) -> pd.DataFrame:
    if _use_bq():
        from . import bq
        return bq.load_p3_bq(company)
    path = find_file(company, "3")
    if path is None:
        log.warning("P3 nao encontrada para %s", company)
        return pd.DataFrame()
    log.info("Lendo P3 (%s): %s", company, os.path.basename(path))
    df = pd.read_excel(path, sheet_name="Plan1", engine="openpyxl")
    cm = col_map(df)
    for key in ("Data Entrada", "Data Emissao"):
        orig = cm.get(key)
        if orig:
            df[orig] = pd.to_datetime(df[orig], errors="coerce")
    # Adiciona coluna _nf normalizada
    numero_col = cm.get("Numero")
    if numero_col:
        df["_nf"] = df[numero_col].apply(norm_nf)
    return df


# ---------------------------------------------------------------------------
# Planilha 2 - Geral
# ---------------------------------------------------------------------------
def load_geral(company: str) -> pd.DataFrame:
    if _use_bq():
        from . import bq
        return bq.load_geral_bq(company)
    path = find_file(company, "2")
    if path is None:
        log.warning("Geral nao encontrada para %s", company)
        return pd.DataFrame()
    sheet = config.SHEET_GERAL[company]
    log.info("Lendo Geral (%s) aba '%s': %s", company, sheet, os.path.basename(path))
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    # Normalizar nomes de colunas: strip + mapear aliases
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]

    # Para Novitah: mapear COD BARRAS -> COD PROD se COD PROD nao existir
    existing = set(asciize(c) for c in df.columns)
    if "COD PROD" not in existing and "COD BARRAS" in existing:
        for c in df.columns:
            if asciize(c) == "COD BARRAS":
                df = df.rename(columns={c: "COD PROD"})
                log.info("Coluna mapeada: COD BARRAS -> COD PROD (%s)", company)
                break

    # Adicionar coluna _nf normalizada
    cm = col_map(df)
    nf_col = cm.get("NF")
    if nf_col:
        df["_nf"] = df[nf_col].apply(norm_nf)

    # Converter LANCAMENTO para datetime
    lanc_col = cm.get("LANCAMENTO")
    if lanc_col:
        df[lanc_col] = pd.to_datetime(df[lanc_col], errors="coerce")

    return df


# ---------------------------------------------------------------------------
# Aba Lancamentos (sempre do arquivo da Alinare)
# ---------------------------------------------------------------------------
def load_lancamentos() -> pd.DataFrame:
    """Le a aba Lancamentos do arquivo P2 da Alinare via openpyxl (preserva texto das celulas)."""
    path = find_file("alinare", "2")
    if path is None:
        log.warning("P2 da Alinare nao encontrado")
        return pd.DataFrame()
    log.info("Lendo aba Lancamentos (Alinare) via openpyxl")
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[4]  # Indice 4 = Lancamentos
    wb.close()
    return ws


def parse_lancamentos(ws) -> list[dict]:
    """Extrai dados da aba Lancamentos em uma lista de dicionarios."""
    pronto_kw = ("programado", "finalizado", "lancado")
    records = []
    header_rows = {4, 15}

    for r in range(1, ws.max_row + 1):
        if r in header_rows:
            continue

        bu = ws.cell(row=r, column=2).value
        data_raw = ws.cell(row=r, column=3).value
        desc = ws.cell(row=r, column=4).value
        embarque = ws.cell(row=r, column=5).value
        mkt_val = ws.cell(row=r, column=6).value
        status = ws.cell(row=r, column=7).value

        bu_str = str(bu).strip().lower() if bu else ""
        if bu_str not in ("alinare", "novitah"):
            continue

        # Classificar status (normalizando acentos)
        s_norm = asciize(str(status)).lower() if status else ""
        is_ready = any(kw in s_norm for kw in pronto_kw)

        # MKT: verificar se contem "entregue"
        mkt_ok = "entregue" in str(mkt_val).lower() if mkt_val else False

        # Data: datetime, "Pendente", ou outro
        if isinstance(data_raw, datetime):
            data_val = data_raw.strftime("%d/%m") if data_raw.year == config.PROXIMO_MES.year and data_raw.month == config.PROXIMO_MES.month else None
            if data_val is None:
                continue  # Nao e do proximo mes
            in_scope = True
        elif str(data_raw).strip().lower() == "pendente":
            data_val = "Sem data"
            in_scope = True
        else:
            continue

        if not in_scope:
            continue

        records.append({
            "empresa": bu_str,
            "data": data_val,
            "descricao": str(desc).strip() if desc else "",
            "embarque": str(embarque).strip() if embarque else "",
            "status_raw": str(status).strip() if status else "",
            "status": "OK" if is_ready else "EM PROCESSO",
            "mkt": "OK" if mkt_ok else "EM PROCESSO",
        })

    return records


def load_lancamentos_bq() -> list[dict]:
    """Tela 3 via BigQuery (bronze.lancamentos). Delega ao adapter bq."""
    from . import bq
    return bq.load_lancamentos_bq()
