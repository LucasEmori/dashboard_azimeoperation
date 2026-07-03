"""Carregamento robusto de planilhas Excel + conversao de datas + cache parquet."""
from __future__ import annotations

import hashlib
import logging
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from . import config

log = logging.getLogger("dashboard.io")

EXCEL_EPOCH = datetime(1899, 12, 30)


# ---------------------------------------------------------------------------
# Conversao de datas
# ---------------------------------------------------------------------------
def excel_to_date(val) -> pd.Timestamp | pd._libs.tslibs.nattype.NaTType:
    """Converte serial Excel ou objeto de data para Timestamp."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return pd.NaT
    if isinstance(val, datetime):
        return pd.Timestamp(val)
    if isinstance(val, date):
        return pd.Timestamp(val)
    if isinstance(val, (pd.Timestamp,)):
        return val
    try:
        n = float(val)
    except (TypeError, ValueError):
        return pd.NaT
    if n <= 0 or n > 200000:
        return pd.NaT
    return pd.Timestamp(EXCEL_EPOCH + timedelta(days=n))


# ---------------------------------------------------------------------------
# Cache
# ---------------------------------------------------------------------------
def _cache_key(path: Path) -> str:
    st = path.stat()
    raw = f"{path}|{st.st_size}|{st.st_mtime_ns}"
    return hashlib.md5(raw.encode()).hexdigest()[:12]


def _read_cache(name: str, key: str) -> pd.DataFrame | None:
    f = config.CACHE_DIR / f"{name}__{key}.pkl"
    if f.exists():
        log.info("cache hit: %s", f.name)
        return pd.read_pickle(f)
    return None


def _write_cache(df: pd.DataFrame, name: str, key: str) -> None:
    config.CACHE_DIR.mkdir(parents=True, exist_ok=True)
    f = config.CACHE_DIR / f"{name}__{key}.pkl"
    df.to_pickle(f)
    log.info("cache escrito: %s", f.name)


# ---------------------------------------------------------------------------
# Planilha 1 - Produtos Lancados
# ---------------------------------------------------------------------------
def load_planilha1(company: str) -> pd.DataFrame:
    path = config.find_file(company, config.FILE_P1)
    if path is None:
        log.warning("Planilha 1 nao encontrada para %s", company)
        return pd.DataFrame()
    key = _cache_key(path)
    cached = _read_cache(f"p1_{company}", key)
    if cached is not None:
        return _clean_planilha1_dates(cached)
    log.info("Lendo Planilha 1 (%s): %s", company, path.name)
    df = pd.read_excel(path, sheet_name="Plan1", engine="openpyxl")
    _write_cache(df, f"p1_{company}", key)
    return _clean_planilha1_dates(df)


def _clean_planilha1_dates(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    for col in ("Data", "Data Lançamento", "Data Virada"):
        if col in df.columns:
            df[col] = df[col].apply(excel_to_date)
    return df


# ---------------------------------------------------------------------------
# Planilha 3 - Notas de Entrada
# ---------------------------------------------------------------------------
def load_planilha3(company: str) -> pd.DataFrame:
    path = config.find_file(company, config.FILE_P3)
    if path is None:
        log.warning("Planilha 3 nao encontrada para %s", company)
        return pd.DataFrame()
    key = _cache_key(path)
    cached = _read_cache(f"p3_{company}", key)
    if cached is not None:
        return _clean_planilha3_dates(cached)
    log.info("Lendo Planilha 3 (%s): %s", company, path.name)
    df = pd.read_excel(path, sheet_name="Plan1", engine="openpyxl")
    _write_cache(df, f"p3_{company}", key)
    return _clean_planilha3_dates(df)


def _clean_planilha3_dates(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    for col in ("Data Emissão", "Data Entrada", "Dt. Quitação", "Data Autor."):
        if col in df.columns:
            df[col] = df[col].apply(excel_to_date)
    return df


# ---------------------------------------------------------------------------
# Planilha 2 - Geral (aba "Geral" ou "BASE GERAL")
# ---------------------------------------------------------------------------
def _normalize_geral_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza nomes de colunas: strip espacos e mapeia alias para nomes canonicos."""
    if df.empty:
        return df
    # 1. Strip espacos de todas as colunas
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    # 2. Mapear alias para nomes canonicos (se o nome canonico nao existir)
    existing = set(df.columns)
    rename = {}
    for canonical, aliases in config.COL_ALIASES.items():
        if canonical not in existing:
            for alias in aliases:
                if alias in existing and alias not in rename:
                    rename[alias] = canonical
                    break
    if rename:
        df = df.rename(columns=rename)
        log.info("Colunas normalizadas: %s", rename)
    return df


def load_geral(company: str) -> pd.DataFrame:
    path = config.find_file(company, config.FILE_P2)
    if path is None:
        log.warning("Planilha Geral nao encontrada para %s", company)
        return pd.DataFrame()
    key = _cache_key(path)
    cached = _read_cache(f"geral_{company}", key)
    if cached is not None:
        return _clean_geral_dates(cached)
    sheet_name = config.SHEET_GERAL_BY_COMPANY.get(company, config.SHEET_GERAL)
    log.info("Lendo aba '%s' (%s): %s [pode demorar...]", sheet_name, company, path.name)
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    df = _normalize_geral_columns(df)
    _write_cache(df, f"geral_{company}", key)
    return _clean_geral_dates(df)


def _clean_geral_dates(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    if "LANÇAMENTO" in df.columns:
        df["LANÇAMENTO"] = df["LANÇAMENTO"].apply(excel_to_date)
    if "Data Último Lancto" in df.columns:
        df["Data Último Lancto"] = df["Data Último Lancto"].apply(excel_to_date)
    return df


# ---------------------------------------------------------------------------
# Planilha 2 - aba "Lancamentos" (com cores das celulas -> MKT)
# ---------------------------------------------------------------------------
def load_lancamentos(company: str) -> pd.DataFrame:
    """Le a aba Lancamentos via openpyxl preservando cores de preenchimento.

    Retorna DataFrame com colunas: empresa_bloco, mes, data, embarque_pedra,
    embarque, mkt_valor, mkt_verde, transferencia, configurador, status_bruto.
    """
    path = config.find_file(company, config.FILE_P2)
    if path is None:
        log.warning("Planilha Geral nao encontrada para %s", company)
        return pd.DataFrame()
    log.info("Lendo aba Lancamentos (%s) via openpyxl", company)
    theme_map = _build_theme_map(path)
    wb = load_workbook(path, data_only=True, read_only=False)
    if config.SHEET_LANCAMENTOS not in wb.sheetnames:
        log.warning("Aba Lancamentos ausente em %s", company)
        return pd.DataFrame()
    ws = wb[config.SHEET_LANCAMENTOS]
    rows = _parse_lancamentos_sheet(ws, theme_map)
    wb.close()
    return rows


def _build_theme_map(path: Path) -> dict[int, tuple[int, int, int]]:
    """Le xl/theme/theme1.xml e mapeia indices de tema openpyxl -> (R,G,B)."""
    import zipfile
    import re as _re
    try:
        with zipfile.ZipFile(path) as z:
            theme = z.read("xl/theme/theme1.xml").decode("utf-8", "ignore")
    except (KeyError, FileNotFoundError):
        return {}
    hexes = _re.findall(r'<a:srgbClr val="([0-9A-Fa-f]{6})"/>', theme)
    if len(hexes) < 10:
        return {}
    # Ordem no XML: dk1, lt1, dk2, lt2, accent1..accent6, hlink, folHlink
    xml_order = hexes[:10]
    # Mapear para indices openpyxl (swap dk1/lt1, dk2/lt2)
    # openpyxl: 0=lt1, 1=dk1, 2=lt2, 3=dk2, 4=accent1..9=accent6
    names = ["dk1", "lt1", "dk2", "lt2", "accent1", "accent2", "accent3",
             "accent4", "accent5", "accent6"]
    xml_map = dict(zip(names, xml_order))
    opx_map = {
        0: xml_map["lt1"], 1: xml_map["dk1"], 2: xml_map["lt2"], 3: xml_map["dk2"],
        4: xml_map["accent1"], 5: xml_map["accent2"], 6: xml_map["accent3"],
        7: xml_map["accent4"], 8: xml_map["accent5"], 9: xml_map["accent6"],
    }
    result = {}
    for idx, hx in opx_map.items():
        result[idx] = (int(hx[0:2], 16), int(hx[2:4], 16), int(hx[4:6], 16))
    return result


def _apply_tint(rgb: tuple[int, int, int], tint: float) -> tuple[int, int, int]:
    """Aplica tint (clarear >0 / escurecer <0) a uma cor RGB."""
    r, g, b = rgb
    if tint > 0:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    elif tint < 0:
        r = int(r * (1 + tint))
        g = int(g * (1 + tint))
        b = int(b * (1 + tint))
    return (max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b)))


def _resolve_fill_rgb(cell, theme_map: dict) -> tuple[int, int, int] | None:
    """Resolve a cor de preenchimento de uma celula para (R,G,B)."""
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    tint = fg.tint or 0.0
    if fg.type == "rgb" and isinstance(fg.rgb, str):
        rgb_str = fg.rgb
        if rgb_str.startswith("00"):
            rgb_str = rgb_str[2:]
        if len(rgb_str) >= 6:
            return _apply_tint(
                (int(rgb_str[0:2], 16), int(rgb_str[2:4], 16), int(rgb_str[4:6], 16)),
                tint,
            )
    if fg.type == "theme" and fg.theme in theme_map:
        return _apply_tint(theme_map[fg.theme], tint)
    if fg.type == "indexed":
        # indexed color 43 = verde, 50 = verde claro, 14 = verde
        green_idx = {43, 50, 14, 4}
        if fg.indexed in green_idx:
            return (0, 255, 0)
    return None


def _is_green(rgb: tuple[int, int, int]) -> bool:
    """Verifica se uma cor RGB e' predominantemente verde."""
    r, g, b = rgb
    return g > 100 and g > r * 1.25 and g > b * 1.25


def _parse_lancamentos_sheet(ws, theme_map: dict | None = None) -> pd.DataFrame:
    if theme_map is None:
        theme_map = {}
    c = config.LANC_COL
    records = []
    for r in range(1, ws.max_row + 1):
        if r in config.LANC_HEADER_ROWS:
            continue
        bu = ws.cell(row=r, column=c["bu"]).value
        data_raw = ws.cell(row=r, column=c["data"]).value
        embarque_pedra = ws.cell(row=r, column=c["embarque_pedra"]).value
        embarque = ws.cell(row=r, column=c["embarque"]).value
        mkt_cell = ws.cell(row=r, column=c["mkt"])
        mkt_valor = mkt_cell.value
        mkt_rgb = _resolve_fill_rgb(mkt_cell, theme_map)
        mkt_verde = _is_green(mkt_rgb) if mkt_rgb else False
        status = ws.cell(row=r, column=c["status"]).value
        transferencia = ws.cell(row=r, column=c["transferencia"]).value

        vals = [bu, data_raw, embarque_pedra, embarque, mkt_valor, status, transferencia]
        if all(v in (None, "") for v in vals):
            continue

        bu_str = str(bu).strip().lower() if bu is not None else ""
        empresa = "novitah" if "novitah" in bu_str else ("alinare" if "alinare" in bu_str else None)

        data = excel_to_date(data_raw)
        records.append({
            "empresa": empresa,
            "bu": bu,
            "data": data,
            "embarque_pedra": embarque_pedra,
            "embarque": embarque,
            "mkt_valor": mkt_valor,
            "mkt_verde": mkt_verde,
            "status": status,
            "transferencia": transferencia,
        })
    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Util: descobrir codigo MARCA da Novitah dinamicamente
# ---------------------------------------------------------------------------
def detect_marca_codes(geral: pd.DataFrame) -> dict[str, set[str]]:
    """Retorna mapeando empresa -> conjunto de codigos MARCA encontrados."""
    if geral.empty or "MARCA" not in geral.columns:
        return {}
    codes = set(geral["MARCA"].dropna().astype(str).str.upper().unique())
    alinare = {c for c in codes if c in config.MARCA_CODES["alinare"] or c.startswith("AL")}
    novitah = {c for c in codes if c in config.MARCA_CODES["novitah"] or c.startswith(("NV", "NO", "NT"))}
    resto = codes - alinare - novitah
    if resto and not novitah:
        novitah = resto
    return {"alinare": alinare, "novitah": novitah, "_todos": codes}
